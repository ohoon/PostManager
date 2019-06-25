from tkinter import filedialog
from tkinter import messagebox
from urllib import request as url
from zipfile import BadZipFile
from bs4 import BeautifulSoup
from selenium import webdriver
from selenium.webdriver.common.alert import Alert
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
from selenium.common.exceptions import TimeoutException
from selenium.common.exceptions import NoSuchWindowException
import time
import tkinter as tk
import tkinter.ttk
import re
import webbrowser
import openpyxl

def result():
    global count
    global rcount
    global _detached
    _detached = set()

    def _add():
        global count
        global rcount

        try:
            s = int(r_firstE.get())
            e = int(r_lastE.get())
            if s > e:
                raise ValueError
            if len(str(s)) != 13 or len(str(e)) != 13:
                raise ValueError
        except:
            messagebox.showinfo("Error!", "올바른 값을 입력해주세요.")
        else:
            for a_pnum in range(s, e + 1):
                rpt = False

                for tv_pnum in totalT.get_children():
                    if str(a_pnum) == str(tv_pnum):
                        messagebox.showinfo("Error!", str(a_pnum) + "는 이미 조회된 등기번호입니다.")
                        rpt = True
                        break
                if not rpt:
                    with url.urlopen("https://service.epost.go.kr/trace.RetrieveDomRigiTraceList.comm?sid1=" + str(
                            a_pnum) + "&displayHeader=N") as a_doc:
                        a_html = a_doc.read().decode()
                        a_soup = BeautifulSoup(a_html, features="lxml")
                        a_data = a_soup.find("table", "table_col").find("tbody").find("tr").find_all("td")
                        a_name = str(a_data[1])[4:-5].split("<br/>")[0]
                        a_date = str(a_data[1])[4:-5].split("<br/>")[1]
                        a_status = re.sub("\(.*\)", '', a_data[3].get_text())
                        s_a_status = ""
                        r_a_status = ""

                        if a_status == "배달완료" or a_status == "교부":
                            try:
                                s_a_data = re.findall("\(수령인:.*\)", str(a_soup))
                                s_a_status = re.sub("\(수령인:|\)", '', s_a_data[0])
                            except:
                                s_a_status = "Error"
                            finally:
                                succT.insert("", "end", text="", values=[a_pnum, a_name, s_a_status, a_date],
                                             iid=a_pnum)
                                totalT.insert("", "end", text="", values=[a_pnum, a_name, s_a_status, a_date],
                                              iid=a_pnum)
                                count += 1
                        elif a_status == "배달준비":
                            retrnT.insert("", "end", text="", values=[a_pnum, a_name, a_status, a_date], iid=a_pnum)
                            totalT.insert("", "end", text="", values=[a_pnum, a_name, a_status, a_date], iid=a_pnum)
                            rcount += 1
                        elif a_status == "" or a_status == "도착":
                            retrnT.insert("", "end", text="", values=[a_pnum, a_name, a_status, a_date], iid=a_pnum)
                            totalT.insert("", "end", text="", values=[a_pnum, a_name, a_status, a_date], iid=a_pnum)
                            rcount += 1
                        else:
                            try:
                                r_a_data = re.findall("<!-- 미배달사유 -->.*\s*.*\s*.*<!-- //미배달사유 -->", str(a_soup))
                                r_a_status = re.sub("<.+?>|\s", '', r_a_data[len(r_a_data) - 1])
                            except:
                                r_a_status = "Error"
                            finally:
                                retrnT.insert("", "end", text="", values=[a_pnum, a_name, r_a_status, a_date],
                                              iid=a_pnum)
                                totalT.insert("", "end", text="", values=[a_pnum, a_name, r_a_status, a_date],
                                              iid=a_pnum)
                                rcount += 1

            resultL.configure(text="배달완료 " + str(count) + "건\t미배달 " + str(rcount) + "건\t총 " + str(count + rcount) + "건")

    def _pop_up_s(event):
        webbrowser.open("https://service.epost.go.kr/trace.RetrieveDomRigiTraceList.comm?sid1=" + str(
            succT.focus()) + "&displayHeader=N")

    def _pop_up_r(event):
        webbrowser.open("https://service.epost.go.kr/trace.RetrieveDomRigiTraceList.comm?sid1=" + str(
            retrnT.focus()) + "&displayHeader=N")

    def _pop_up_t(event):
        webbrowser.open("https://service.epost.go.kr/trace.RetrieveDomRigiTraceList.comm?sid1=" + str(
            totalT.focus()) + "&displayHeader=N")

    def _search(sv, tv, c, e):
        global _detached
        children = list(_detached) + list(tv.get_children())
        _detached = set()
        query = str(e.get())
        i_r = -1
        if c.get() == "등기번호":
            option = "#1"
        elif c.get() == "성명":
            option = "#2"
        elif c.get() == "수령인" or c.get() == "처리현황" or c.get() == "수령인/처리현황":
            option = "#3"
        elif c.get() == "수령일자" or c.get() == "처리일자" or c.get() == "수령/처리일":
            option = "#4"
        else:
            return

        for i_d in children:
            text = str(tv.set(i_d, option))
            if query in text:
                i_r += 1
                tv.reattach(i_d, '', i_r)
            else:
                _detached.add(i_d)
                tv.detach(i_d)

    def _sort_column(tv, col, reverse):
        letter = ""

        try:
            letter = [(int(tv.set(k, col)), k) for k in tv.get_children('')]
            letter.sort(reverse=reverse)
        except:
            letter = [(tv.set(k, col), k) for k in tv.get_children('')]
            letter.sort(reverse=reverse)
        finally:
            # rearrange items in sorted positions
            for index, (val, k) in enumerate(letter):
                tv.move(k, '', index)

            # reverse sort next time
            tv.heading(col, command=lambda: _sort_column(tv, col, not reverse))

    def database():
        global excel

        def _browse():
            global excel

            filename = filedialog.askopenfilename(initialdir="./", title="Choose Your File",
                                                  filetypes=(("Excel files", "*.xlsx"), ("All files", "*.*")))
            if filename:
                if fileE.get():
                    excel.close()
                fileE.configure(state="normal")
                fileE.delete(0, len(fileE.get()))
                fileE.insert(0, filename)
                fileE.configure(state="readonly")
                try:
                    excel = openpyxl.load_workbook(fileE.get())
                except BadZipFile:
                    messagebox.showinfo(title="파일 열기 오류", message="접근할 수 없는 파일입니다.")
                    fileE.configure(state="normal")
                    fileE.delete(0, len(fileE.get()))
                    fileE.configure(state="readonly")
                    sheetC.configure(values=[])
                    sheetC.set("")
                except FileNotFoundError:
                    messagebox.showinfo(title="파일 열기 오류", message="존재하지 않는 파일입니다.")
                    fileE.configure(state="normal")
                    fileE.delete(0, len(fileE.get()))
                    fileE.configure(state="readonly")
                    sheetC.configure(values=[])
                    sheetC.set("")
                else:
                    sheetC.configure(values=excel.sheetnames)
                    sheetC.current(0)

        def _nmcol(sv, sc, ncc):
            global excel
            global cache

            if sc.get():
                sheet = excel[sc.get()]
                cache = []
                cols = sheet.columns
                for col in cols:
                    for cell in col:
                        if cell.value:
                            cache += [cell]
                            break
                if cache:
                    v = []
                    for c in cache:
                        v += [c.value]
                    ncc.configure(values=v)
                    ncc.current(0)
            else:
                ncc.configure(values=[])
                ncc.set("")

        def unmask():
            global excel
            global cache

            def _pop_up_um(event):
                if str(um_resultT.set(um_resultT.focus(), "#3")):
                    webbrowser.open("https://service.epost.go.kr/trace.RetrieveDomRigiTraceList.comm?sid1=" + str(
                        um_resultT.set(um_resultT.focus(), "#3")) + "&displayHeader=N")

            if not fileE.get():
                messagebox.showinfo("Error!", "명단 파일을 첨부하세요.")
            elif not sheetC.get():
                messagebox.showinfo("Error!", "해당 시트를 찾을 수 없습니다.")
            elif not nmcolC.get():
                messagebox.showinfo("Error!", "해당 열을 불러올 수 없습니다.")
            else:
                um_resultC = tk.Toplevel(dataC)
                um_resultC.title("마스킹 해제 결과")
                um_resultC.geometry("400x307+465+300")
                um_resultC.resizable(False, False)
                um_resultC.lift()
                um_resultT = tkinter.ttk.Treeview(um_resultC, columns=["", "", "", "", ""], height=14)
                um_resultT.grid(column=0, columnspan=2, row=0)
                um_resultT.yview()
                um_resultT["show"] = "headings"
                um_resultT.column("#1", width=30, anchor="center")
                um_resultT.heading("#1", text="#", anchor="center",
                                   command=lambda: _sort_column(um_resultT, "#1", False))
                um_resultT.column("#2", width=45, anchor="center")
                um_resultT.heading("#2", text="성명", anchor="center",
                                   command=lambda: _sort_column(um_resultT, "#2", False))
                um_resultT.column("#3", width=120, anchor="center")
                um_resultT.heading("#3", text="등기번호", anchor="center",
                                   command=lambda: _sort_column(um_resultT, "#3", False))
                um_resultT.column("#4", width=110, anchor="center")
                um_resultT.heading("#4", text="수령인/처리현황", anchor="center",
                                   command=lambda: _sort_column(um_resultT, "#4", False))
                um_resultT.column("#5", width=92, anchor="center")
                um_resultT.heading("#5", text="비고", anchor="center",
                                   command=lambda: _sort_column(um_resultT, "#5", False))
                um_resultT.bind('<<TreeviewOpen>>', _pop_up_um)
                um_resultT.tag_configure('success', background='#B7F0B1')
                um_resultT.tag_configure('fail', background='#F15F5F')
                um_resultT.tag_configure('alert', background='#F2CB61')
                sheet = excel[sheetC.get()]

                index = 0
                s_cnt = 0
                a_cnt = 0
                f_cnt = 0
                n_data = []
                s_list = []
                nm_col = cache[nmcolC.current()].column
                nm_row = cache[nmcolC.current()].row
                if not n_iv.get():
                    nm_row += 1
                for i in range(nm_row, sheet.max_row + 1):
                    n_data += [sheet[i][nm_col - 1].value]
                for tv_nm in totalT.get_children():
                    index += 1
                    swt = False
                    driver = webdriver.Ie('IEDriverServer.exe')
                    webdriver.IeOptions()
                    driver.get(
                        'https://service.epost.go.kr/trace.RetrieveDomRigiTraceList.comm?sid1=' + tv_nm + '&displayHeader=N')
                    main_window_handle = driver.current_window_handle
                    driver.execute_script("fnPopupMaskingSolv()")
                    for handle in driver.window_handles:
                        if handle != main_window_handle:
                            driver.switch_to.window(handle)
                            break
                    for ex_nm in n_data:
                        if ex_nm:
                            msk_nm = ex_nm[0:1] + "*" + ex_nm[2:]
                            if totalT.set(tv_nm, "#2") == msk_nm:
                                driver.execute_script('SetExpCookie("TraceCnt","0",10);')
                                elem = driver.find_element_by_name('senderNm_masking')
                                elem.send_keys("치")
                                elem = driver.find_element_by_name('receverNm_masking')
                                elem.send_keys(ex_nm[1])
                                driver.execute_script("return verifyNms(event);")
                                try:
                                    WebDriverWait(driver, 5).until(EC.alert_is_present())
                                    Alert(driver).accept()
                                    continue
                                except NoSuchWindowException or TimeoutException:
                                    um_pnum = tv_nm
                                    um_status = totalT.set(tv_nm, "#3")
                                    swt = True
                                    if n_data.count(ex_nm) > 1:
                                        um_resultT.insert("", "end", text="",
                                                          values=[index, ex_nm, um_pnum, um_status, "동명이인 존재"],
                                                          iid=index, tags=('alert',))
                                        a_cnt += 1
                                    elif ex_nm in s_list:
                                        for um_list in um_resultT.get_children():
                                            if um_resultT.set(um_list, "#2") == ex_nm:
                                                um_resultT.set(um_list, "#5", "수신인 중복")
                                                um_resultT.item(um_list, tags=('alert',))
                                                s_cnt -= 1
                                                a_cnt += 1
                                        um_resultT.insert("", "end", text="",
                                                          values=[index, ex_nm, um_pnum, um_status, "수신인 중복"],
                                                          iid=index, tags=('alert',))
                                        a_cnt += 1
                                    else:
                                        um_resultT.insert("", "end", text="",
                                                          values=[index, ex_nm, um_pnum, um_status, ""], iid=index,
                                                          tags=('success',))
                                        s_cnt += 1
                                        s_list += [ex_nm]

                                    driver.switch_to.window(main_window_handle)
                                    if p_iv.get():
                                        driver.execute_script("window.print();")
                                        try:
                                            WebDriverWait(driver, 5).until(EC.alert_is_present())
                                            Alert(driver).accept()
                                            time.sleep(5)
                                            driver.close()
                                            break
                                        except TimeoutException:
                                            print("print error")
                                    else:
                                        driver.close()
                                        break
                    if not swt:
                        um_resultT.insert("", "end", text="",
                                          values=[index, totalT.set(tv_nm, "#2"), totalT.set(tv_nm, "#1"),
                                                  totalT.set(tv_nm, "#3"), "성명 불일치"], iid=index, tags=('fail',))
                        f_cnt += 1
                        driver.close()
                        driver.switch_to.window(main_window_handle)
                        driver.close()

                messagebox.showinfo("작업 완료", "성공: " + str(s_cnt) + " 경고: " + str(a_cnt) + " 실패: " + str(f_cnt))

                if x_iv.get():
                    new_workbook = openpyxl.Workbook()
                    new_sheet = new_workbook.active
                    new_sheet.cell(row=1, column=1).value = '#'
                    new_sheet.cell(row=1, column=2).value = '성명'
                    new_sheet.cell(row=1, column=3).value = '등기번호'
                    new_sheet.cell(row=1, column=4).value = '수령인/처리현황'
                    new_sheet.cell(row=1, column=5).value = '비고'
                    now = time.localtime()
                    idx = 2
                    for tv_nm in um_resultT.get_children():
                        new_sheet.cell(row=idx, column=1).value = um_resultT.set(tv_nm, "#1")
                        new_sheet.cell(row=idx, column=2).value = um_resultT.set(tv_nm, "#2")
                        new_sheet.cell(row=idx, column=3).value = um_resultT.set(tv_nm, "#3")
                        new_sheet.cell(row=idx, column=4).value = um_resultT.set(tv_nm, "#4")
                        new_sheet.cell(row=idx, column=5).value = um_resultT.set(tv_nm, "#5")
                        idx += 1

                    new_workbook.save(str(now.tm_year) + str(now.tm_mon) + str(now.tm_mday) + str(now.tm_hour) + str(now.tm_min) + str(now.tm_sec) + ".xlsx")

        dataC = tk.Toplevel(resultC, padx=5, pady=5)
        dataC.title("데이터 불러오기")
        dataC.geometry("250x142+200+300")
        dataC.resizable(False, False)
        dataC.lift()

        fileL = tk.Label(dataC, text="명단 파일", font="함초롬돋움 10")
        fileL.grid(column=0, row=0)
        fileE = tk.Entry(dataC, width=15, state="readonly", readonlybackground="white")
        fileE.grid(column=1, columnspan=2, row=0, padx=3)
        fileB = tk.Button(dataC, text="불러오기", command=_browse)
        fileB.grid(column=3, row=0)

        h_sv = tk.StringVar()
        h_sv.trace("w", lambda nm, index, mode, sv=h_sv: _nmcol(sv, sheetC, nmcolC))
        sheetL = tk.Label(dataC, text="시트 선택", font="함초롬돋움 10")
        sheetL.grid(column=0, row=1)
        sheetC = tkinter.ttk.Combobox(dataC, width=13, state="readonly", textvariable=h_sv)
        sheetC.grid(column=1, columnspan=2, row=1, padx=3)

        n_iv = tk.IntVar()
        nmcolL = tk.Label(dataC, text="이름 열", font="함초롬돋움 10")
        nmcolL.grid(column=0, row=2)
        nmcolC = tkinter.ttk.Combobox(dataC, width=10, state="readonly")
        nmcolC.grid(column=1, row=2, padx=3, sticky="w")
        nmcolCB = tk.Checkbutton(dataC, text="첫 행 포함", font="함초롬돋움 8", variable=n_iv)
        nmcolCB.grid(column=2, columnspan=2, row=2)

        p_iv = tk.IntVar()
        x_iv = tk.IntVar()
        optionL = tk.Label(dataC, text="기타 기능", font="함초롬돋움 10")
        optionL.grid(column=0, row=3)
        printCB = tk.Checkbutton(dataC, text="프린트", font="함초롬돋움 8", variable=p_iv)
        printCB.grid(column=1, row=3)
        printCB = tk.Checkbutton(dataC, text="to xslx", font="함초롬돋움 8", variable=x_iv)
        printCB.grid(column=2, columnspan=2, row=3)
        dataB = tk.Button(dataC, text="마스킹 해제", padx=30, pady=5, command=unmask)
        dataB.grid(column=0, columnspan=4, row=4)

    try:
        fst = int(firstE.get())
        lst = int(lastE.get())

        if fst > lst:
            raise ValueError
        if len(str(fst)) != 13 or len(str(lst)) != 13:
            raise ValueError
    except:
        messagebox.showinfo("Error!", "올바른 값을 입력해주세요.")
    else:
        mainC.iconify()
        resultC = tk.Toplevel(mainC)
        resultC.title("조회 결과")
        resultC.geometry("1146x472+100+300")
        resultC.resizable(0, 0)
        resultC.lift()

        succF = tk.Frame(resultC, bd=3)
        succF.grid(column=0, row=0)
        succL = tk.Label(succF, text="배달완료", font="함초롬돋움 12")
        succL.grid(column=0, columnspan=3, row=0)

        s_scroll = tk.Scrollbar(succF)
        succT = tkinter.ttk.Treeview(succF, columns=["", "", "", ""], height=18, yscrollcommand=s_scroll.set)
        succT.grid(column=0, columnspan=2, row=1)
        succT["show"] = "headings"
        succT.column("#1", width=120, anchor="center")
        succT.heading("#1", text="등기번호", anchor="center", command=lambda: _sort_column(succT, "#1", False))
        succT.column("#2", width=45, anchor="center")
        succT.heading("#2", text="성명", anchor="center", command=lambda: _sort_column(succT, "#2", False))
        succT.column("#3", width=110, anchor="center")
        succT.heading("#3", text="수령인", anchor="center", command=lambda: _sort_column(succT, "#3", False))
        succT.column("#4", width=80, anchor="center")
        succT.heading("#4", text="수령일자", anchor="center", command=lambda: _sort_column(succT, "#4", False))
        succT.bind('<<TreeviewOpen>>', _pop_up_s)
        s_scroll.grid(column=2, row=1, sticky="nws")
        s_scroll["command"] = succT.yview

        s_srchC = tkinter.ttk.Combobox(succF, width=10, values=["등기번호", "성명", "수령인", "수령일자"], state="readonly")
        s_srchC.grid(column=0, row=2, sticky="e")
        s_srchC.set("목록 선택")
        s_sv = tk.StringVar()
        s_sv.trace("w", lambda nm, index, mode, sv=s_sv: _search(sv, succT, s_srchC, s_srchE))
        s_srchE = tk.Entry(succF, width=37, textvariable=s_sv)
        s_srchE.grid(column=1, row=2, sticky="w")

        retrnF = tk.Frame(resultC, bd=3)
        retrnF.grid(column=1, row=0)
        retrnL = tk.Label(retrnF, text="미배달", font="함초롬돋움 12")
        retrnL.grid(column=0, columnspan=3, row=0)

        r_scroll = tk.Scrollbar(retrnF)
        retrnT = tkinter.ttk.Treeview(retrnF, columns=["", "", "", ""], height=18, yscrollcommand=r_scroll.set)
        retrnT.grid(column=0, columnspan=2, row=1)
        retrnT.yview()
        retrnT["show"] = "headings"
        retrnT.column("#1", width=120, anchor="center")
        retrnT.heading("#1", text="등기번호", anchor="center", command=lambda: _sort_column(retrnT, "#1", False))
        retrnT.column("#2", width=45, anchor="center")
        retrnT.heading("#2", text="성명", anchor="center", command=lambda: _sort_column(retrnT, "#2", False))
        retrnT.column("#3", width=110, anchor="center")
        retrnT.heading("#3", text="처리현황", anchor="center", command=lambda: _sort_column(retrnT, "#3", False))
        retrnT.column("#4", width=80, anchor="center")
        retrnT.heading("#4", text="처리일자", anchor="center", command=lambda: _sort_column(retrnT, "#4", False))
        retrnT.bind('<<TreeviewOpen>>', _pop_up_r)
        r_scroll.grid(column=2, row=1, sticky="nws")
        r_scroll["command"] = retrnT.yview

        r_srchC = tkinter.ttk.Combobox(retrnF, width=10, values=["등기번호", "성명", "처리현황", "처리일자"], state="readonly")
        r_srchC.grid(column=0, row=2, sticky="e")
        r_srchC.set("목록 선택")
        r_sv = tk.StringVar()
        r_sv.trace("w", lambda nm, index, mode, sv=r_sv: _search(sv, retrnT, r_srchC, r_srchE))
        r_srchE = tk.Entry(retrnF, width=37, textvariable=r_sv)
        r_srchE.grid(column=1, row=2, sticky="w")

        totalF = tk.Frame(resultC, bd=3)
        totalF.grid(column=2, row=0)
        totalL = tk.Label(totalF, text="통합", font="함초롬돋움 12")
        totalL.grid(column=0, columnspan=3, row=0)

        t_scroll = tk.Scrollbar(totalF)
        totalT = tkinter.ttk.Treeview(totalF, columns=["", "", "", ""], height=18, yscrollcommand=t_scroll.set)
        totalT.grid(column=0, columnspan=2, row=1)
        totalT.yview()
        totalT["show"] = "headings"
        totalT.column("#1", width=120, anchor="center")
        totalT.heading("#1", text="등기번호", anchor="center", command=lambda: _sort_column(totalT, "#1", False))
        totalT.column("#2", width=45, anchor="center")
        totalT.heading("#2", text="성명", anchor="center", command=lambda: _sort_column(totalT, "#2", False))
        totalT.column("#3", width=110, anchor="center")
        totalT.heading("#3", text="수령인/처리현황", anchor="center", command=lambda: _sort_column(totalT, "#3", False))
        totalT.column("#4", width=80, anchor="center")
        totalT.heading("#4", text="수령/처리일", anchor="center", command=lambda: _sort_column(totalT, "#4", False))
        totalT.bind('<<TreeviewOpen>>', _pop_up_t)
        t_scroll.grid(column=2, row=1, sticky="nws")
        t_scroll["command"] = totalT.yview

        t_srchC = tkinter.ttk.Combobox(totalF, width=10, values=["등기번호", "성명", "수령인/처리현황", "수령/처리일"], state="readonly")
        t_srchC.grid(column=0, row=2, sticky="e")
        t_srchC.set("목록 선택")
        t_sv = tk.StringVar()
        t_sv.trace("w", lambda nm, index, mode, sv=t_sv: _search(sv, totalT, t_srchC, t_srchE))
        t_srchE = tk.Entry(totalF, width=37, textvariable=t_sv)
        t_srchE.grid(column=1, row=2, sticky="w")

        count = 0
        rcount = 0
        for pnum in range(fst, lst + 1):
            with url.urlopen("https://service.epost.go.kr/trace.RetrieveDomRigiTraceList.comm?sid1=" + str(
                    pnum) + "&displayHeader=N") as doc:
                html = doc.read().decode()
                soup = BeautifulSoup(html, features="lxml")
                data = soup.find("table", "table_col").find("tbody").find("tr").find_all("td")
                name = str(data[1])[4:-5].split("<br/>")[0]
                date = str(data[1])[4:-5].split("<br/>")[1]
                status = re.sub("\(.*\)", '', data[3].get_text())

                if status == "배달완료" or status == "교부":
                    try:
                        s_data = re.findall("\(수령인:.*\)", str(soup))
                        s_status = re.sub("\(수령인:|\)", '', s_data[0])
                    except:
                        s_status = "error"
                    finally:
                        succT.insert("", "end", text="", values=[pnum, name, s_status, date], iid=pnum)
                        totalT.insert("", "end", text="", values=[pnum, name, s_status, date], iid=pnum)
                        count += 1
                elif status == "배달준비":
                    retrnT.insert("", "end", text="", values=[pnum, name, status, date], iid=pnum)
                    totalT.insert("", "end", text="", values=[pnum, name, status, date], iid=pnum)
                    rcount += 1
                elif status == "" or status == "도착":
                    retrnT.insert("", "end", text="", values=[pnum, name, status, date], iid=pnum)
                    totalT.insert("", "end", text="", values=[pnum, name, status, date], iid=pnum)
                    rcount += 1
                elif status == "미배달":
                    try:
                        r_data = re.findall("<!-- 미배달사유 -->.*\s*.*\s*.*<!-- //미배달사유 -->", str(soup))
                        r_status = re.sub("<.+?>|\s", '', r_data[len(r_data) - 1])
                    except:
                        r_status = "error"
                    finally:
                        retrnT.insert("", "end", text="", values=[pnum, name, r_status, date], iid=pnum)
                        totalT.insert("", "end", text="", values=[pnum, name, r_status, date], iid=pnum)
                        rcount += 1

        resultF = tk.Frame(resultC)
        resultF.grid(column=0, columnspan=3, row=1)
        resultL = tk.Label(resultF,
                           text="배달완료 " + str(count) + "건\t미배달 " + str(rcount) + "건\t총 " + str(count + rcount) + "건",
                           font="함초롬돋움 10", anchor="center", padx=120)
        resultL.grid(column=0, row=0)
        resultB = tk.Button(resultF, text="데이터 불러오기", width=24, font="함초롬돋움 8", command=database)
        resultB.grid(column=1, row=0, padx=35)
        r_firstE = tk.Entry(resultF)
        r_firstE.grid(column=2, row=0, padx=3)
        r_Label = tk.Label(resultF, text="~")
        r_Label.grid(column=3, row=0)
        r_lastE = tk.Entry(resultF)
        r_lastE.grid(column=4, row=0, padx=3)
        result2B = tk.Button(resultF, text="이어서 조회", width=12, font="함초롬돋움 9", command=_add)
        result2B.grid(column=5, row=0)


mainC = tk.Tk()
mainC.title("등기번호 조회기")
mainC.geometry("230x129+500+400")
mainC.resizable(0, 0)

mainL = tk.Label(mainC, text="등기번호(13자리)", font="함초롬돋움 15")
mainL.grid(column=0, columnspan=2, row=0)

firstL = tk.Label(mainC, text="시작 번호\t", font="함초롬돋움 12")
firstL.grid(column=0, row=1)
firstE = tk.Entry(mainC)
firstE.grid(column=1, row=1)

lastL = tk.Label(mainC, text=" 끝 번호\t", font="함초롬돋움 12")
lastL.grid(column=0, row=2)
lastE = tk.Entry(mainC)
lastE.grid(column=1, row=2)

mainB = tk.Button(mainC, text="조회하기", width=15, font="함초롬돋움 13", command=result)
mainB.grid(column=0, columnspan=2, row=3)
mainC.mainloop()
